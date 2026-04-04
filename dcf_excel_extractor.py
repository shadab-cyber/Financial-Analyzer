# dcf_excel_extractor.py
"""
Extract CFO and CAPEX from a Screener.in Excel export for DCF valuation.

Handles two Screener.in export layouts:
  A) Named CF sheet with actual values  (older exports)
  B) Named CF sheet with formula refs   (newer exports) → fall back to Data Sheet
"""
import re as _re
import pandas as pd
import numpy as np
import openpyxl


def fmt(v):
    if pd.isna(v):
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


# ─── Year detection ───────────────────────────────────────────────────────────

def _is_year_value(val):
    """Return True if val looks like a fiscal year label from Screener.in."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    if hasattr(val, 'year'):          # datetime / Timestamp
        return True
    if isinstance(val, (int, float)) and 2000 <= float(val) <= 2100:
        return True
    if isinstance(val, str):
        s = val.strip()
        # "Mar 2024", "March 2024", "Mar-2024"
        if _re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-]20\d{2}$', s, _re.I):
            return True
        # ISO "2024-03-31"
        if _re.match(r'^20\d{2}[\-\/]\d{2}[\-\/]\d{2}$', s):
            return True
        # bare year "2024"
        if _re.match(r'^20\d{2}$', s):
            return True
        # Indian FY "2023-24" or "2023–24" (en-dash)
        if _re.match(r'^20\d{2}[\-\u2013]\d{2,4}$', s):
            return True
        # "FY2024", "FY 2024", "FY24"
        if _re.match(r'^FY\s*(\d{2}|\d{4})$', s, _re.I):
            return True
        try:
            parsed = pd.to_datetime(s, dayfirst=False)
            if 2000 <= parsed.year <= 2100:
                return True
        except Exception:
            pass
    return False


def _year_label(val):
    """Convert any year value to a readable 'Mar YYYY' string."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if hasattr(val, 'year'):
        return val.strftime('Mar %Y')
    if isinstance(val, (int, float)) and 2000 <= float(val) <= 2100:
        return f'Mar {int(val)}'
    if isinstance(val, str):
        s = val.strip()
        if _re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-]20\d{2}$', s, _re.I):
            return s.replace('-', ' ')
        # Indian FY "2023-24" → "Mar 2024"
        m = _re.match(r'^(20\d{2})[\-\u2013](\d{2,4})$', s)
        if m:
            return f'Mar {int(m.group(1)) + 1}'
        # "FY2024" → "Mar 2024"
        m = _re.match(r'^FY\s*(20\d{2})$', s, _re.I)
        if m:
            return f'Mar {m.group(1)}'
        # "FY24" → "Mar 2024"
        m = _re.match(r'^FY\s*(\d{2})$', s, _re.I)
        if m:
            return f'Mar 20{m.group(1)}'
        try:
            return pd.to_datetime(s).strftime('Mar %Y')
        except Exception:
            return s
    return str(val)


def _find_year_row(df, start_col=1, min_dates=2):
    """Return the row index with the most fiscal year columns, or None."""
    best_row, best_count = None, 0
    for row_idx in range(df.shape[0]):       # scan ALL rows, not just first 40
        count = sum(
            1 for col in range(start_col, df.shape[1])
            if _is_year_value(df.iloc[row_idx, col])
        )
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row if best_count >= min_dates else None


def _extract_year_cols(df, year_row, start_col=1):
    """
    Return (years_list, col_indices_list) for all year-valued cells
    in year_row from start_col onwards.
    Skips NaN gaps and non-year tokens like 'TTM'.
    """
    years, cols = [], []
    for col in range(start_col, df.shape[1]):
        val = df.iloc[year_row, col]
        if not _is_year_value(val):
            continue
        label = _year_label(val)
        if label:
            years.append(label)
            cols.append(col)
    return years, cols


# ─── Sheet reading ────────────────────────────────────────────────────────────

def _is_formula_string(val):
    return isinstance(val, str) and val.startswith('=')


def _sheet_has_real_values(df, year_row, start_col=1):
    """Return True if the year row contains actual values (not formula strings)."""
    for col in range(start_col, df.shape[1]):
        val = df.iloc[year_row, col]
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            if not _is_formula_string(val):
                return True
    return False


def _read_cf_from_data_sheet(filepath):
    """
    Newer Screener.in exports store formulas in the named CF sheet but keep
    real values in 'Data Sheet'.  This function extracts just the Cash Flow
    section from that sheet.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if 'Data Sheet' not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb['Data Sheet']
        all_rows = list(ws.iter_rows(values_only=True))

        # Find "CASH FLOW:" section header (case-insensitive)
        start_idx = None
        for i, row in enumerate(all_rows):
            if row and row[0] and str(row[0]).strip().upper().startswith('CASH FLOW'):
                start_idx = i
                break

        if start_idx is None:
            return pd.DataFrame()

        # Collect rows until we hit the next major section or too many blanks
        result_rows = []
        blank_streak = 0
        for row in all_rows[start_idx:]:
            if all(v is None for v in row):
                blank_streak += 1
                if blank_streak > 3:
                    break
            else:
                blank_streak = 0
                result_rows.append(row)

        return pd.DataFrame(result_rows) if result_rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _read_cf_sheet(filepath):
    """Try all known Screener CF sheet names; fall back to Data Sheet."""
    for name in ('Cash Flow Data', 'Cash Flow Statement', 'Cash Flows',
                 'Cashflow', 'Cash flow', 'Cash Flow'):
        try:
            df = pd.read_excel(filepath, engine='openpyxl',
                               sheet_name=name, header=None)
            if df.empty:
                continue
            year_row = _find_year_row(df)
            if year_row is not None and _sheet_has_real_values(df, year_row):
                return df          # real values found — use this sheet
        except Exception:
            continue

    # Named sheet was formula-only or missing → pull from Data Sheet
    return _read_cf_from_data_sheet(filepath)


# ─── Row label matching ───────────────────────────────────────────────────────

def find_row(df, label):
    """Find row by exact then partial label match (col 0)."""
    if df.empty:
        return None
    label_l = label.strip().lower()
    for i in range(df.shape[0]):
        val = df.iloc[i, 0]
        if isinstance(val, str) and val.strip().lower() == label_l:
            return i
    for i in range(df.shape[0]):
        val = df.iloc[i, 0]
        if isinstance(val, str) and label_l in val.strip().lower():
            return i
    return None


CFO_LABELS = [
    'Cash from Operating Activity',
    'Cash from Operating Activities',
    'Net Cash from Operating Activities',
    'Cash Generated from Operations',
    'Net cash from operating',
    'Operating Cash Flow',
    'CFO',
]

CAPEX_LABELS = [
    'Fixed Assets Purchased',
    'Purchase of Fixed Assets',
    'Capital Expenditure',
    'Capex',
    'Purchase of Property, Plant',
    'Addition to Fixed Assets',
    'Net Fixed Assets Purchased',
    # Fallback: use investing CF as CAPEX proxy when no line-item detail
    'Cash from Investing Activity',
    'Cash from Investing Activities',
    'Net Cash from Investing Activities',
]


def _find_any_row(df, labels):
    for label in labels:
        r = find_row(df, label)
        if r is not None:
            return r, label
    return None, None


# ─── Main extractor ───────────────────────────────────────────────────────────

def extract_dcf_from_excel(filepath):
    """
    Read a Screener.in Excel file and return:
    {
        'years':        ['Mar 2019', ...],
        'cfo':          [1234.5, ...],
        'capex':        [300.1, ...],   # always positive
        'fcf':          [934.4, ...],
        'capex_source': 'Fixed Assets Purchased' | 'Cash from Investing Activity',
        'source':       'excel',
    }
    Raises ValueError with a helpful message if data cannot be found.
    """
    cf_df = _read_cf_sheet(filepath)

    if cf_df.empty:
        raise ValueError(
            'No Cash Flow sheet found in this Excel file.\n'
            'Please upload a Screener.in Excel export — go to any company page on '
            'Screener.in → click "Export to Excel" at the top.'
        )

    YEAR_ROW = _find_year_row(cf_df)

    if YEAR_ROW is None:
        # Build diagnostic sample
        sample = []
        for r in range(min(5, cf_df.shape[0])):
            row_vals = [str(cf_df.iloc[r, c]) for c in range(1, min(6, cf_df.shape[1]))]
            sample.append(f'  Row {r}: {row_vals}')
        raise ValueError(
            'Could not detect year columns in the Excel file. '
            'Make sure this is a Screener.in export.\n'
            'Sample cell values:\n' + '\n'.join(sample)
        )

    years, year_cols = _extract_year_cols(cf_df, YEAR_ROW, start_col=1)
    n = len(years)

    if n == 0:
        raise ValueError(
            'Could not detect year columns in the Excel file. '
            'Make sure this is a Screener.in export.'
        )

    def row_values(row_idx):
        return [fmt(cf_df.iloc[row_idx, col]) for col in year_cols]

    # ── CFO ──────────────────────────────────────────────────────────────────
    cfo_row, _ = _find_any_row(cf_df, CFO_LABELS)
    if cfo_row is None:
        raise ValueError(
            'Could not find Cash from Operating Activities in the Excel.\n'
            'Expected one of: ' + ', '.join(f'"{l}"' for l in CFO_LABELS[:4]) + '\n'
            'Make sure you are uploading a Screener.in Excel export.'
        )
    cfo = row_values(cfo_row)

    # ── CAPEX ────────────────────────────────────────────────────────────────
    capex_row, capex_label = _find_any_row(cf_df, CAPEX_LABELS)
    if capex_row is None:
        raise ValueError(
            'Could not find Capital Expenditure / Fixed Assets Purchased in the Excel.\n'
            'Expected one of: ' + ', '.join(f'"{l}"' for l in CAPEX_LABELS[:4]) + '\n'
            'Make sure you are uploading a Screener.in Excel export.'
        )
    capex_raw  = row_values(capex_row)
    capex      = [abs(v) for v in capex_raw]   # always positive
    is_proxy   = 'Investing' in (capex_label or '')

    # ── FCF ──────────────────────────────────────────────────────────────────
    fcf = [round(cfo[i] - capex[i], 2) for i in range(n)]

    # Strip trailing all-zero years (no data yet)
    while len(years) > 2 and cfo[-1] == 0 and capex[-1] == 0:
        years.pop(); cfo.pop(); capex.pop(); fcf.pop()

    return {
        'years':        years,
        'cfo':          [round(v, 2) for v in cfo],
        'capex':        [round(v, 2) for v in capex],
        'fcf':          fcf,
        'capex_source': capex_label or 'Unknown',
        'capex_is_proxy': is_proxy,
        'source':       'excel',
    }


# ─── WACC / EBITDA helpers ───────────────────────────────────────────────────

def _read_bs_pl_sheet(filepath):
    """Read Balance Sheet & P&L data, with Data Sheet fallback."""
    for name in ('Balance Sheet & P&L', 'Profit & Loss', 'Balance Sheet'):
        try:
            df = pd.read_excel(filepath, engine='openpyxl',
                               sheet_name=name, header=None)
            if not df.empty:
                yr = _find_year_row(df)
                if yr is not None and _sheet_has_real_values(df, yr):
                    return df
        except Exception:
            continue

    # Fallback: Data Sheet P&L section
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if 'Data Sheet' not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb['Data Sheet']
        all_rows = list(ws.iter_rows(values_only=True))
        start_idx = None
        for i, row in enumerate(all_rows):
            if row and row[0] and str(row[0]).strip().upper().startswith('PROFIT'):
                start_idx = i
                break
        if start_idx is None:
            return pd.DataFrame()
        result, blanks = [], 0
        for row in all_rows[start_idx:]:
            if all(v is None for v in row):
                blanks += 1
                if blanks > 3:
                    break
            else:
                blanks = 0
                result.append(row)
        return pd.DataFrame(result) if result else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def extract_wacc_inputs(filepath):
    result = {'interest': None, 'debt': None, 'tax_rate': None, 'equity': None}
    try:
        bs_df = _read_bs_pl_sheet(filepath)
        if bs_df.empty:
            return result

        YEAR_ROW  = _find_year_row(bs_df) or 0
        START_COL = 1

        def latest(label):
            r = find_row(bs_df, label)
            if r is None:
                return None
            for col in range(bs_df.shape[1] - 1, START_COL - 1, -1):
                v = fmt(bs_df.iloc[r, col])
                if v != 0:
                    return v
            return None

        result['interest'] = latest('Interest')
        result['debt']     = latest('Borrowings')

        r_tax = find_row(bs_df, 'Tax')
        r_pbt = find_row(bs_df, 'Profit before tax')
        if r_tax is not None and r_pbt is not None:
            for col in range(bs_df.shape[1] - 1, START_COL - 1, -1):
                tax = fmt(bs_df.iloc[r_tax, col])
                pbt = fmt(bs_df.iloc[r_pbt, col])
                if pbt and tax and pbt > 0:
                    result['tax_rate'] = round(tax / pbt * 100, 1)
                    break

        eq  = latest('Equity Share Capital') or 0
        res = latest('Reserves')             or 0
        if eq + res > 0:
            result['equity'] = round(eq + res, 2)
    except Exception:
        pass
    return result


def extract_ebitda(filepath):
    try:
        bs_df = _read_bs_pl_sheet(filepath)
        if bs_df.empty:
            return None, None

        YEAR_ROW  = _find_year_row(bs_df) or 0
        START_COL = 1

        def latest_val(label):
            r = find_row(bs_df, label)
            if r is None:
                return None
            for col in range(bs_df.shape[1] - 1, START_COL - 1, -1):
                v = fmt(bs_df.iloc[r, col])
                if v != 0:
                    return v
            return None

        def latest_year():
            for col in range(bs_df.shape[1] - 1, START_COL - 1, -1):
                val = bs_df.iloc[YEAR_ROW, col]
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    return _year_label(val) or str(val)
            return None

        op_profit = latest_val('Operating Profit') or latest_val('EBIT') or 0
        dep       = latest_val('Depreciation') or 0
        ebitda    = op_profit + dep
        return (round(ebitda, 2) if ebitda != 0 else None), latest_year()
    except Exception:
        return None, None


def extract_latest_fcf(filepath):
    try:
        data = extract_dcf_from_excel(filepath)
        if not data['fcf'] or not data['years']:
            return None, None
        return data['fcf'][-1], data['years'][-1]
    except Exception:
        return None, None
