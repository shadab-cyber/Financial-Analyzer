#!/usr/bin/env python3
"""
financialanalyzer.py
Reads Screener.in Excel exports — 'Data Sheet' tab.

Structure of Data Sheet:
  Row with 'PROFIT & LOSS'  → annual P&L rows follow
  Row with 'BALANCE SHEET'  → balance sheet rows follow
  Row with 'CASH FLOW:'     → cash flow rows follow
  Each section has a 'Report Date' row then data rows.
  Columns B..K = up to 10 years of data.
"""

from typing import List, Optional, Dict
import datetime


# ══════════════════════════════════════════════════════════════════════════════
# Excel reader
# ══════════════════════════════════════════════════════════════════════════════

def _clean(v) -> Optional[float]:
    """Convert cell value to float, return None if not a number."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        import re
        s = v.replace(',', '').replace('%', '').strip()
        s = re.sub(r'[^\d.\-]', '', s)
        try:
            return float(s) if s else None
        except ValueError:
            return None
    return None


def _year_label(v) -> str:
    """Convert a date cell to 'Mar-24' style label."""
    if isinstance(v, datetime.datetime):
        return v.strftime('%b-%y')
    return str(v) if v else ''


def read_screener_excel(filepath: str) -> dict:
    """
    Parse Screener.in Excel export — Data Sheet tab.
    Sections: PROFIT & LOSS | Quarters | BALANCE SHEET | CASH FLOW
    Returns dict with keys: years, income, balance, cashflow
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl not installed. Add openpyxl to requirements.txt")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the Data Sheet (case-insensitive)
    sheet_name = None
    for name in wb.sheetnames:
        if 'data' in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError(f"No Data Sheet found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    n_rows = len(rows)

    # ── Find ALL section boundaries ───────────────────────────────────────────
    pl_start = quarters_start = bs_start = cf_start = price_start = None
    for i, row in enumerate(rows):
        label = str(row[0]).strip().upper() if row[0] else ''
        if 'PROFIT' in label and 'LOSS' in label and pl_start is None:
            pl_start = i
        elif label == 'QUARTERS' and quarters_start is None:
            quarters_start = i
        elif label == 'BALANCE SHEET' and bs_start is None:
            bs_start = i
        elif 'CASH FLOW' in label and cf_start is None:
            cf_start = i
        elif label in ('PRICE:', 'PRICE') and price_start is None:
            price_start = i

    # ── Section boundaries (end = next section start) ────────────────────────
    # P&L annual data ends at Quarters section
    pl_end       = quarters_start if quarters_start else (bs_start if bs_start else n_rows)
    # Quarters section skipped entirely
    bs_end        = cf_start if cf_start else n_rows
    # Cash flow ends at PRICE: row
    cf_end        = price_start if price_start else n_rows

    def _parse_section(start: int, end: int) -> Dict[str, List]:
        """
        Parse a section of rows. First 'Report Date' row = year labels.
        Stops at the next section header.
        """
        section: Dict[str, List] = {}
        for r in rows[start + 1: end]:
            label = str(r[0]).strip() if r[0] else ''
            if not label:
                continue
            # Stop if we hit another section header
            upper = label.upper()
            if any(h in upper for h in ['BALANCE SHEET', 'CASH FLOW', 'PRICE:', 'DERIVED', 'QUARTERS']):
                break

            vals = list(r[1:])  # columns B..K

            if label == 'Report Date':
                years = [_year_label(v) for v in vals if v is not None]
                section['_years'] = years
                section['_n']     = len(years)
                continue

            n   = section.get('_n', len(vals))
            nums = [_clean(vals[i]) if i < len(vals) else None for i in range(n)]
            section[label] = nums

        return section

    pl_data = _parse_section(pl_start, pl_end)  if pl_start is not None else {}
    bs_data = _parse_section(bs_start, bs_end)  if bs_start is not None else {}
    cf_data = _parse_section(cf_start, cf_end)  if cf_start is not None else {}

    years = (pl_data.get('_years') or bs_data.get('_years') or cf_data.get('_years') or [])

    return {'years': years, 'income': pl_data, 'balance': bs_data, 'cashflow': cf_data}


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def _get(section: dict, *keys) -> List:
    """Return first matching key's list, else empty list."""
    for k in keys:
        for sk in section:
            if sk.lower().strip() == k.lower().strip():
                return [v for v in section[sk] if v is not None] or section[sk]
    return []


def cagr(vals: List) -> Optional[float]:
    vals = [v for v in vals if v is not None]
    if len(vals) < 2 or vals[-1] <= 0 or vals[0] <= 0:
        return None
    n = len(vals) - 1
    return round(((vals[-1] / vals[0]) ** (1 / n) - 1) * 100, 2)


def avg_growth(vals: List) -> Optional[float]:
    vals = [v for v in vals if v is not None]
    if len(vals) < 2:
        return None
    growths = []
    for i in range(1, len(vals)):
        if vals[i - 1] not in (None, 0):
            growths.append((vals[i] - vals[i - 1]) / abs(vals[i - 1]) * 100)
    return round(sum(growths) / len(growths), 2) if growths else None


def avg(lst: List) -> Optional[float]:
    vals = [v for v in lst if v is not None]
    return round(sum(vals) / len(vals), 2) if vals else None


def ratio_yearly(num: List, den: List) -> List:
    n = max(len(num), len(den))
    out = []
    for i in range(n):
        a = num[i] if i < len(num) else None
        b = den[i] if i < len(den) else None
        out.append(round(a / b, 2) if (a is not None and b not in (None, 0)) else None)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# Analyzers
# ══════════════════════════════════════════════════════════════════════════════

def analyze_income(inc: dict, years: List[str]) -> dict:
    revenue    = _get(inc, 'Sales', 'Revenue', 'Net Sales')
    net_profit = _get(inc, 'Net profit', 'Net Profit', 'PAT')
    interest   = _get(inc, 'Interest', 'Finance Costs')
    tax        = _get(inc, 'Tax', 'Tax Expense')
    depreciation = _get(inc, 'Depreciation', 'Depreciation & Amortisation')
    employee   = _get(inc, 'Employee Cost', 'Employee Benefit Expenses')
    raw_mat    = _get(inc, 'Raw Material Cost', 'Cost of Materials Consumed', 'COGS')
    other_inc  = _get(inc, 'Other Income')
    pbt        = _get(inc, 'Profit before tax', 'PBT')

    # EBITDA = Net Profit + Tax + Interest + Depreciation
    n = max(len(net_profit), len(tax), len(interest), len(depreciation))
    ebitda = []
    for i in range(n):
        np = net_profit[i]  if i < len(net_profit)   else 0
        tx = tax[i]         if i < len(tax)           else 0
        it = interest[i]    if i < len(interest)      else 0
        dp = depreciation[i] if i < len(depreciation) else 0
        if any(v is not None for v in [np, tx, it, dp]):
            ebitda.append(round((np or 0)+(tx or 0)+(it or 0)+(dp or 0), 2))

    def pct(a, b): return round(a / b * 100, 2) if (a is not None and b not in (None, 0)) else None

    # Latest year metrics
    rev0  = revenue[0]    if revenue    else None
    np0   = net_profit[0] if net_profit else None
    ebit0 = ebitda[0]     if ebitda     else None

    opm = pct(ebit0, rev0)
    npm = pct(np0, rev0)

    return {
        'Statement':                   'Income Statement',
        'Years':                       years,
        'Revenue (All Years)':         revenue,
        'Net Profit (All Years)':      net_profit,
        'EBITDA (All Years)':          ebitda,
        'PBT (All Years)':             pbt,
        'Depreciation (All Years)':    depreciation,
        'Interest (All Years)':        interest,
        'Tax (All Years)':             tax,
        'Employee Cost (All Years)':   employee,
        'Raw Material Cost (All Yrs)': raw_mat,
        'Other Income (All Years)':    other_inc,
        '5Y Revenue CAGR (%)':         cagr(revenue[-6:] if len(revenue) > 6 else revenue),
        '5Y Net Profit CAGR (%)':      cagr(net_profit[-6:] if len(net_profit) > 6 else net_profit),
        '5Y EBITDA CAGR (%)':          cagr(ebitda[-6:] if len(ebitda) > 6 else ebitda),
        'Avg Revenue Growth (%)':      avg_growth(revenue),
        'Avg Net Profit Growth (%)':   avg_growth(net_profit),
        'EBITDA Margin (%)':           opm,
        'Net Profit Margin (%)':       npm,
    }


def analyze_balance(bs: dict, inc: dict, years: List[str]) -> dict:
    equity_cap = _get(bs, 'Equity Share Capital', 'Share Capital')
    reserves   = _get(bs, 'Reserves', 'Reserves and Surplus')
    borrowings = _get(bs, 'Borrowings', 'Total Borrowings')
    other_liab = _get(bs, 'Other Liabilities')
    total      = _get(bs, 'Total')
    net_block  = _get(bs, 'Net Block', 'Fixed Assets')
    investments= _get(bs, 'Investments')
    other_assets=_get(bs, 'Other Assets')
    receivables= _get(bs, 'Receivables', 'Debtors')
    inventory  = _get(bs, 'Inventory', 'Inventories')
    cash       = _get(bs, 'Cash & Bank', 'Cash and Cash Equivalents')

    # Equity = Share Capital + Reserves
    n = max(len(equity_cap), len(reserves))
    equity = []
    for i in range(n):
        ec = equity_cap[i] if i < len(equity_cap) else 0
        rs = reserves[i]   if i < len(reserves)   else 0
        equity.append(round((ec or 0) + (rs or 0), 2))

    # Total debt
    debt = borrowings  # from screener this is total borrowings

    de_ratio   = ratio_yearly(debt, equity)
    # Current ratio approximation: (Cash + Receivables + Inventory) / Other Liabilities
    cur_assets = []
    for i in range(max(len(cash), len(receivables), len(inventory))):
        c  = cash[i]        if i < len(cash)        else 0
        r  = receivables[i] if i < len(receivables) else 0
        iv = inventory[i]   if i < len(inventory)   else 0
        cur_assets.append(round((c or 0)+(r or 0)+(iv or 0), 2))
    current_ratio = ratio_yearly(cur_assets, other_liab)

    net_profit = _get(inc, 'Net profit', 'Net Profit', 'PAT')
    interest   = _get(inc, 'Interest', 'Finance Costs')
    tax        = _get(inc, 'Tax', 'Tax Expense')

    roa = ratio_yearly([v * 100 if v else v for v in net_profit], total)
    roe = ratio_yearly([v * 100 if v else v for v in net_profit], equity)

    # Interest coverage = EBIT / Interest
    n2 = max(len(net_profit), len(tax), len(interest))
    ebit_list = []
    for i in range(n2):
        np = net_profit[i] if i < len(net_profit) else 0
        tx = tax[i]        if i < len(tax)        else 0
        it = interest[i]   if i < len(interest)   else 0
        ebit_list.append((np or 0)+(tx or 0)+(it or 0))
    ic = ratio_yearly(ebit_list, interest)

    return {
        'Statement':                                 'Balance Sheet',
        'Years':                                     years,
        'Equity Share Capital (All Years)':          equity_cap,
        'Reserves (All Years)':                      reserves,
        'Total Equity (All Years)':                  equity,
        'Borrowings (All Years)':                    borrowings,
        'Total Assets (All Years)':                  total,
        'Net Block (All Years)':                     net_block,
        'Investments (All Years)':                   investments,
        'Cash & Bank (All Years)':                   cash,
        'Receivables (All Years)':                   receivables,
        'Inventory (All Years)':                     inventory,
        'Approx Current Assets (All Years)':         cur_assets,
        'Debt-to-Equity (All Years)':                de_ratio,
        'Approx Current Ratio (All Years)':          current_ratio,
        'Return on Assets ROA % (All Years)':        roa,
        'Return on Equity ROE % (All Years)':        roe,
        'Interest Coverage Ratio (All Years)':       ic,
        '5Y Avg D/E':                                avg(de_ratio[-6:]),
        '5Y Avg ROA (%)':                            avg(roa[-6:]),
        '5Y Avg ROE (%)':                            avg(roe[-6:]),
        '5Y Avg Current Ratio':                      avg(current_ratio[-6:]),
    }


def analyze_cashflow(cf: dict, bs: dict, inc: dict, years: List[str]) -> dict:
    cfo   = _get(cf, 'Cash from Operating Activity', 'CFO', 'Operating Cash Flow')
    cfi   = _get(cf, 'Cash from Investing Activity', 'CFI', 'Investing Cash Flow')
    cff   = _get(cf, 'Cash from Financing Activity', 'CFF', 'Financing Cash Flow')
    netchg= _get(cf, 'Net Cash Flow', 'Net Change in Cash')

    revenue    = _get(inc, 'Sales', 'Revenue')
    net_profit = _get(inc, 'Net profit', 'Net Profit')
    total      = _get(bs, 'Total')
    equity_cap = _get(bs, 'Equity Share Capital')
    reserves   = _get(bs, 'Reserves')
    equity = [round((equity_cap[i] or 0)+(reserves[i] or 0), 2)
              for i in range(max(len(equity_cap), len(reserves)))]

    # Free Cash Flow = CFO + CFI (CFI is usually negative for capex)
    fcf = []
    for i in range(min(len(cfo), len(cfi))):
        o = cfo[i]; fi = cfi[i]
        fcf.append(round((o or 0)+(fi or 0), 2) if (o is not None or fi is not None) else None)

    # Earnings quality = CFO / Net Profit
    eq_ratio = ratio_yearly(cfo, net_profit)
    # CFO margin = CFO / Revenue
    cfo_margin = ratio_yearly([v * 100 if v else v for v in cfo], revenue)
    # Cash ROE = CFO / Equity * 100
    cash_roe = ratio_yearly([v * 100 if v else v for v in cfo], equity)
    # Cash ROA = CFO / Total Assets * 100
    cash_roa = ratio_yearly([v * 100 if v else v for v in cfo], total)

    return {
        'Statement':                              'Cash Flow',
        'Years':                                  years,
        'Operating Cash Flow (All Years)':        cfo,
        'Investing Cash Flow (All Years)':        cfi,
        'Financing Cash Flow (All Years)':        cff,
        'Free Cash Flow (All Years)':             fcf,
        'Net Change in Cash (All Years)':         netchg,
        'CFO Margin % (All Years)':               cfo_margin,
        'Cash ROE % (All Years)':                 cash_roe,
        'Cash ROA % (All Years)':                 cash_roa,
        'Earnings Quality Ratio (All Years)':     eq_ratio,
        '5Y CFO CAGR (%)':                        cagr(cfo[-6:] if len(cfo) > 6 else cfo),
        '5Y Avg CFO Margin (%)':                  avg(cfo_margin[-6:]),
        '5Y Avg Earnings Quality':                avg(eq_ratio[-6:]),
    }


def analyze_excel(filepath: str) -> dict:
    """
    Main entry point. Pass path to Screener.in Excel file.
    Returns full analysis dict ready for jsonify().
    """
    data = read_screener_excel(filepath)
    years   = data['years']
    inc     = data['income']
    bs      = data['balance']
    cf      = data['cashflow']

    result = {}

    if inc:
        try:
            result['Income Statement'] = analyze_income(inc, years)
        except Exception as e:
            result['Income Statement'] = {'error': str(e)}

    if bs:
        try:
            result['Balance Sheet'] = analyze_balance(bs, inc, years)
        except Exception as e:
            result['Balance Sheet'] = {'error': str(e)}

    if cf:
        try:
            result['Cash Flow'] = analyze_cashflow(cf, bs, inc, years)
        except Exception as e:
            result['Cash Flow'] = {'error': str(e)}

    return result


# ── Legacy PDF stubs — kept so app.py import doesn't break ──────────────────
def pdf_to_text(pdf_path: str): return []
def is_image_pdf(pdf_path: str) -> bool: return False
def detect_type(lines, filename=''): return None
def extract_income_series(lines): return {}
def extract_balance_series(lines): return {}
def extract_cashflow_series(lines): return {}


if __name__ == '__main__':
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else 'Quick_Heal_Tech.xlsx'
    result = analyze_excel(path)
    print(json.dumps(result, indent=2, default=str))
